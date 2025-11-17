import pandas as pd
import random
import re
from openpyxl import Workbook


class DataPreparer:
    def __init__(self, file_path):
        self.file_path = file_path
        self.sheet_names = []
        self.all_sheets = {}
        self.courses = {}
        self.instructors = []
        self.lab_types = set()
        self._load_data()

    def _load_data(self):
        self.sheet_names = pd.ExcelFile(self.file_path).sheet_names

        for sheet_name in self.sheet_names:
            self.all_sheets[sheet_name] = pd.read_excel(
                self.file_path, sheet_name=sheet_name
            )

        if "Professors" in self.sheet_names:
            self.sheet_names.remove("Professors")
        self.sheet_names.sort()

        for sheet_name in self.sheet_names:
            df = self.all_sheets[sheet_name]
            if not all(
                col in df.columns
                for col in ["Branch", "Course", "Course Instructor", "Course Type"]
            ):
                continue
            if "Lab Type" not in df.columns:
                df["Lab Type"] = pd.NA
            if "Combined With" not in df.columns:
                df["Combined With"] = pd.NA

            for _, row in df.iterrows():
                branch = str(row["Branch"]).strip().lower()
                course = str(row["Course"]).strip()
                instructor = str(row["Course Instructor"]).strip()
                course_type_raw = str(row["Course Type"]).strip()
                combined_with = str(row["Combined With"]).strip()
                lab_type = str(row["Lab Type"]).strip()

                lec_match = re.search(r"(\d+)L", course_type_raw.upper())
                lab_match = re.search(r"(\d+)H", course_type_raw.upper())
                lectures = int(lec_match.group(1)) if lec_match else 0
                lab_hours = int(lab_match.group(1)) if lab_match else 0
                if lectures == 0 and lab_hours == 0:
                    continue

                if instructor and instructor not in self.instructors:
                    self.instructors.append(instructor)
                if lab_type and lab_type != "nan":
                    self.lab_types.add(lab_type)

                key = f"{sheet_name.lower()}_{branch}"
                if key not in self.courses:
                    self.courses[key] = []
                self.courses[key].append(
                    [course, lectures, lab_hours, instructor, combined_with, lab_type]
                )

    def get_courses(self):
        return self.courses

    def get_instructors(self):
        return self.instructors

    def get_lab_types(self):
        return list(self.lab_types)


class TimeTableSolver:
    def __init__(self, instructors, courses, lab_types):
        self.instructors = instructors
        self.courses = courses
        self.lab_types = lab_types
        self.all_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        self.all_times = ["9", "10", "11", "12", "2", "3", "4", "5"]

        self.CONTINUOUS_SLOTS = {
            1: [("9",), ("10",), ("11",), ("12",), ("2",), ("3",), ("4",), ("5",)],
            2: [("9", "10"), ("11", "12"), ("2", "3"), ("4", "5")],
            3: [("9", "10", "11"), ("2", "3", "4")],
        }

        base_tt = pd.DataFrame(
            index=self.all_days, columns=self.all_times, dtype="string"
        ).fillna("0")
        self.instructor_tt = {i: base_tt.copy() for i in self.instructors}
        self.sem_tt = {sem: base_tt.copy() for sem in self.courses.keys()}
        self.lab_tt = {lab: base_tt.copy() for lab in self.lab_types}
        self.combined_classes = {}

    def shuffle_courses(self, lst):
        random.shuffle(lst)
        return lst

    def _apply_slots(self, slots, tag, instructor, sem, other_sem, lab_type):
        # For student timetable, use simple tag
        for d, t in slots:
            self.sem_tt[sem].loc[d, t] = tag
            if other_sem:
                self.sem_tt[other_sem].loc[d, t] = tag

        # For instructor and lab timetables, include semester and branch info
        # Extract course name and type (LAB/LEC)
        course_name = tag.split(" (")[0]
        course_type = tag.split("-")[-1] if "-" in tag else ""

        # Format: COURSE-SEM-BRANCH-TYPE (e.g., CS111C-3-CSE-LAB)
        sem_parts = sem.split("_")
        sem_num = sem_parts[0].replace("sem", "").strip()
        branch = sem_parts[1].upper() if len(sem_parts) > 1 else ""

        if branch:
            detailed_tag = f"{course_name}-{sem_num}-{branch}-{course_type}"
        else:
            detailed_tag = f"{course_name}-{sem_num}-{course_type}"

        for d, t in slots:
            self.instructor_tt[instructor].loc[d, t] = detailed_tag
            if lab_type:
                self.lab_tt[lab_type].loc[d, t] = detailed_tag

    def _find_slot(self, hours, tag, instructor, sem, other_sem, lab_type):
        if hours == 0:
            return [], "OK"
        for day in random.sample(self.all_days, len(self.all_days)):
            for block in self.CONTINUOUS_SLOTS.get(hours, []):
                if all(
                    self.instructor_tt[instructor].loc[day, t] == "0"
                    and self.sem_tt[sem].loc[day, t] == "0"
                    for t in block
                ):
                    slots = [(day, t) for t in block]
                    return slots, "OK"
        return [], "Failed"

    def solve(self, sem):
        result = []
        for c, lec, lab, inst, comb, lab_type in self.shuffle_courses(
            self.courses[sem]
        ):
            tag = f"{c} ({sem})"
            other_sem = (
                sem[: sem.find("_")] + "_" + comb.lower()
                if comb and comb != "nan"
                else None
            )

            lab_slots, lab_status = self._find_slot(
                lab, tag + "-LAB", inst, sem, other_sem, lab_type
            )
            if lab_status == "OK":
                self._apply_slots(
                    lab_slots, tag + "-LAB", inst, sem, other_sem, lab_type
                )

            lec_slots, lec_status = self._find_slot(
                lec, tag + "-LEC", inst, sem, other_sem, None
            )
            if lec_status == "OK":
                self._apply_slots(lec_slots, tag + "-LEC", inst, sem, other_sem, None)

            all_slots = lab_slots + lec_slots
            slot_str = (
                ", ".join([f"{d[:3]}-{t}" for d, t in all_slots])
                if all_slots
                else "---"
            )
            status = "OK" if lab_status == "OK" and lec_status == "OK" else "Failed"
            result.append([c, inst, status, slot_str])
        return result


def run_and_save_solver(input_path, output_path):
    dp = DataPreparer(input_path)
    courses = dp.get_courses()
    instructors = dp.get_instructors()
    labs = dp.get_lab_types()

    solver = TimeTableSolver(instructors, courses, labs)
    all_solution = {sem: solver.solve(sem) for sem in courses.keys()}

    wb = Workbook()
    del wb["Sheet"]

    # --- MODIFIED: Add prefixes to sheet names for categorization ---

    # 1. Save report sheets (frontend will ignore these)
    for sem, sol in all_solution.items():
        ws = wb.create_sheet(title=f"REPORT-{sem}"[:30])
        ws.append(["Course", "Instructor", "Status", "Assigned Slots"])
        for r in sol:
            ws.append(r)

    # 2. Save Student (Semester) timetable sheets
    for sem, tt in solver.sem_tt.items():
        # Use the prefix React expects: TT-Sem-
        ws = wb.create_sheet(title=f"TT-Sem-{sem}"[:30])
        ws.append(["Day/Time"] + list(tt.columns))
        for d in tt.index:
            ws.append([d] + list(tt.loc[d]))

    # 3. Save Professor (Instructor) timetable sheets
    for inst, tt in solver.instructor_tt.items():
        # Use the prefix React expects: TT-Inst-
        ws = wb.create_sheet(title=f"TT-Inst-{inst}"[:30])
        ws.append(["Day/Time"] + list(tt.columns))
        for d in tt.index:
            ws.append([d] + list(tt.loc[d]))

    # 4. Save Lab Room timetable sheets
    for lab, tt in solver.lab_tt.items():
        # Use the prefix React expects: TT-Lab-
        ws = wb.create_sheet(title=f"TT-Lab-{lab}"[:30])
        ws.append(["Day/Time"] + list(tt.columns))
        for d in tt.index:
            ws.append([d] + list(tt.loc[d]))

    # --- END MODIFICATION ---

    wb.save(output_path)
    print(f"✅ Saved to {output_path}")
